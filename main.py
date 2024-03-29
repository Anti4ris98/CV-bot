import telebot
import webbrowser
from telebot import types
import openpyxl
from pathlib import Path


bot = telebot.TeleBot('6896862671:AAE9GWFSnReoGNHF5Q0cvi03iPDRjqNWvM4')

@bot.message_handler(commands=['start'])
def main(message):
    if message.from_user.language_code == 'uk':
         bot.send_message(message.chat.id, f"Доброго дня {message.from_user.first_name}{message.from_user.last_name}, Мене звати Ілля. То з чого почнемо знайомство? Напишіть слово <b>Допомога</b> і я розкажу що я могу вам розказати", parse_mode='html')

    elif message.from_user.language_code == 'en':
        bot.send_message(message.chat.id, f"Hello {message.from_user.first_name}{message.from_user.last_name}, my name is Illia Lobachov, let`s start our acquaintance. Just write <b>Assistance</b> and I will describe you what I can what can I tell about myself", parse_mode='html')

@bot.message_handler(commands=['summary'])
def summary(message):
    file_path = Path(__file__).parent / "infoeng.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.worksheets[0]
        for row in sheet.iter_rows(values_only=True):
            row_data = ', '.join(map(str, row))
            bot.send_message(message.chat.id, row_data)
    except Exception as e:
        bot.send_message(message.chat.id, f"An error occurred: {str(e)}")


@bot.message_handler(commands=['academic'])
def summary(message):
    file_path = Path(__file__).parent / "infoeng.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.worksheets[1]
        for row in sheet.iter_rows(values_only=True):
            row_data = ', '.join(map(str, row))
            bot.send_message(message.chat.id, row_data)
    except Exception as e:
        bot.send_message(message.chat.id, f"An error occurred: {str(e)}")


@bot.message_handler(commands=['experience'])
def summary(message):
    file_path = Path(__file__).parent / "infoeng.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.worksheets[2]
        for row in sheet.iter_rows(values_only=True):
            row_data = ', '.join(map(str, row))
            bot.send_message(message.chat.id, row_data)
    except Exception as e:
        bot.send_message(message.chat.id, f"An error occurred: {str(e)}")


@bot.message_handler(commands=['contact'])
def summary(message):
    file_path = Path(__file__).parent / "infoeng.xlsx"
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.worksheets[3]
        for row in sheet.iter_rows(values_only=True):
            row_data = ', '.join(map(str, row))
            bot.send_message(message.chat.id, row_data)
    except Exception as e:
        bot.send_message(message.chat.id, f"An error occurred: {str(e)}")


@bot.message_handler(commands=['info'])
def main(message):
    bot.send_message(message.chat.id, message)
    
@bot.message_handler(commands=['git'])
def site(message):
    webbrowser.open('https://github.com/Anti4ris98')

@bot.message_handler(commands=['linkedin'])
def site(message):
    webbrowser.open('https://www.linkedin.com/in/ilya-lobachov-927873188/')

@bot.message_handler(commands=['storage'])
def site(message):
    webbrowser.open('https://drive.google.com/drive/u/0/folders/1e7p3FlCashIPPDcno7_iMhC-huw7ogFL?q=sharedwith:public%20parent:1e7p3FlCashIPPDcno7_iMhC-huw7ogFL')


@bot.message_handler()
def help(message):
    ukr_comands = {'<em>самарі</em>': 'розповім що вмію',
                    '<em>навчання</em>':  'розповім де вчився',
                    '<em>досвід</em>': 'розповім свою професійну історію',
                    '<em>контакти</em>':  'як зі мною звязатись',
                    '<em>скачати</em>': 'Підготую документ для скачування'
    }
    eng_comands = {'<em>/summary</em>': 'describe my abbilities',
                   '<em>/academic</em>': 'tell about my academic history',
                   '<em>/experience</em>': 'tell you my сareer пrowth',
                   '<em>/contact</em>': 'how to contack with me',
                   '<em>/export</em>': 'prepare document for download'
    }
        
    if message.text.lower() == 'допомога':
        bot.send_message(message.chat.id, '<b>За допомогою цих фраз я...</b>', parse_mode='html')
        for command, description in ukr_comands.items():
            bot.send_message(message.chat.id, f"<b>{command}</b>: {description}", parse_mode='html')
    
    elif message.text.lower() == 'assistance':
        bot.send_message(message.chat.id, '<b>By using this comands I will ...</b>', parse_mode='html')
        for command, description in eng_comands.items():
            bot.send_message(message.chat.id, f"<b>{command}</b>: {description}", parse_mode='html')



    elif message.text.lower() == 'самарі':
        file_path = Path(__file__).parent / "infoua.xlsx"
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.worksheets[0]
            all_rows_data = []  # Створюємо пустий список для зберігання даних з кожного рядка

            for row in sheet.iter_rows(values_only=True):
                row_data = list(row)  # Конвертуємо кортеж у список
                all_rows_data.append(row_data)  # Додаємо список до загального списку

    # Відправляємо загальний список з даними у Telegram
            for row_data in all_rows_data:
                bot.send_message(message.chat.id, ', '.join(map(str, row_data)))
        
        except Exception as e:
            bot.send_message(message.chat.id, f"Виникла помилка: {str(e)}")

    elif message.text.lower() == 'навчання':
        file_path = Path(__file__).parent / "infoua.xlsx"
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.worksheets[1]
            for row in sheet.iter_rows(values_only=True):
                row_data = ', '.join(map(str, row))
                bot.send_message(message.chat.id, row_data)
        except Exception as e:
            bot.send_message(message.chat.id, f"An error occurred: {str(e)}")

    elif message.text.lower() == 'досвід':
        file_path = Path(__file__).parent / "infoua.xlsx"
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.worksheets[2]
            for row in sheet.iter_rows(values_only=True):
                row_data = ', '.join(map(str, row))
                bot.send_message(message.chat.id, row_data)
        except Exception as e:
            bot.send_message(message.chat.id, f"An error occurred: {str(e)}")

    elif message.text.lower() == 'контакти':
        file_path = Path(__file__).parent / "infoua.xlsx"
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.worksheets[3]
            for row in sheet.iter_rows(values_only=True):
                row_data = ', '.join(map(str, row))
                bot.send_message(message.chat.id, row_data)
        except Exception as e:
            bot.send_message(message.chat.id, f"An error occurred: {str(e)}")






bot.polling(non_stop=True)